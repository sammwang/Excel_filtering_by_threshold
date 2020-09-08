import openpyxl as xl;
import sys
from openpyxl.styles import Font

#===============Variables for the script=================#

targetExcel="Memory_Highlighted_Output.xlsx"
threshold=80
thresholdName="Avg Util(%)"
thresholdRow=2

#========================================================#

def autoextract(inputfile,outputfile):
	# opening the source excel file
	filename = inputfile
	wb1 = xl.load_workbook(filename)

	# opening the destination excel file
	filename1 = outputfile
	wb2 = xl.load_workbook(filename1)
	ws2 = wb2["Output"]
	ws3 = wb2["Counter"]

	sheets = wb1.sheetnames
	print("The sheets for processing include:", sheets)
	x=len(sheets)
	for z in range(x):
		ws1 = wb1.worksheets[z]

		# calculate total number of rows and
		# columns in the source excel file
		mr = ws1.max_row
		mc = ws1.max_column

		#Initialize a variable for checking the column position of reference value.
		referVal=999
		for a in range(1, mc + 1):
			if ws1.cell(row=thresholdRow, column=a).value == thresholdName:
				referVal=a
				break

		if referVal == 999:
				print("ERROR: No Threshold Name cell was found.")
				sys.exit(1)

		#Set header
		ws2['A1'] = "Network"
		ws2['B1'] = "Device IP Address"
		ws2['C1'] = "Device Name"
		ws2['D1'] = "Product Series"
		ws2['E1'] = "Memory Pool"
		ws2['F1'] = "Min Util(%)"
		ws2['G1'] = "Max Util(%)"
		ws2['H1'] = "Avg Util(%)"

		#Set font style
		ws2['A1'].font = Font(bold=True)
		ws2['B1'].font = Font(bold=True)
		ws2['C1'].font = Font(bold=True)
		ws2['D1'].font = Font(bold=True)
		ws2['E1'].font = Font(bold=True)
		ws2['F1'].font = Font(bold=True)
		ws2['G1'].font = Font(bold=True)
		ws2['H1'].font = Font(bold=True)

		#Counter to count the number of devices in a network.
		count=0

		# copying the cell values from source
		# excel file to destination excel file
		for i in range(1, mr + 1):
			if i>2:
				if float(ws1.cell(row=i, column=referVal).value) >= threshold:
					lastrow = len(ws2['A']) # Check the last row of Column A for appending.
					count+=1
					ws2.cell(row=lastrow+1, column=1).value = ws1.title
					for j in range(1, mc + 1):
						# reading cell value from source excel file
						c = ws1.cell(row=i, column=j)

						# writing the read value to destination excel file
						ws2.cell(row=lastrow+1, column=j+1).value = c.value

		ws3['A1'] = "Network"
		ws3['B1'] = "Number of Devices"
		lastrowCounterSheet = len(ws3['A'])  # Check the last row of Column A for appending.
		ws3.cell(row=lastrowCounterSheet + 1, column=1).value = ws1.title
		ws3.cell(row=lastrowCounterSheet + 1, column=2).value = count

	lastrowCounterSheet = len(ws3['A'])
	ws3.cell(row=lastrowCounterSheet + 1, column=1).value = "Total"

	# Calculate the total and paste in a temp Cell, copy it to the original Total Cell.
	ws3.cell(row=lastrowCounterSheet + 1, column=2).value = "=SUM(B1:B" + str(lastrowCounterSheet) + ")"
	#print(ws3.cell(row=lastrowCounterSheet + 1, column=2).value)

	# saving the destination excel file
	wb2.save(str(filename1))

def clearsheet(outputfile):
	wb2 = xl.load_workbook(outputfile)
	ws2 = wb2["Output"]
	for row in ws2['A1:Z999']:
		for cell in row:
			cell.value = None

	ws3 = wb2["Counter"]
	for row in ws3['A1:Z99']:
		for cell in row:
			cell.value = None

	wb2.save(str(outputfile))

if __name__ == '__main__':
	if len(sys.argv) < 2:
		print("ERROR: Please enter a valid source filename.")
		sys.exit(1)

	try:
		print("Source filename: %s" % (sys.argv[1]))
		sourceExcel = sys.argv[1]
		clearsheet(targetExcel)
		autoextract(sourceExcel,targetExcel)
		print("The program has been completed. Please check the output file:", targetExcel)
	except:
		print("Unexpected error:", sys.exc_info()[0])
		raise